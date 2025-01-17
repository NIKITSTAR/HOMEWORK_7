import io
import os
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv


def test_create_archive():
    if not os.path.exists('resources'):
        os.mkdir('resources')
    with zipfile.ZipFile('resources/archive.zip', 'w') as create:
        for file in ['testfile1.pdf', 'testfile2.xlsx', 'testfile3.csv']:
            if os.path.exists(file):
                create.write(file, os.path.basename(file))
            else:
                print(f'Файл {file} не найден в корне')


def test_read_and_check():
    with zipfile.ZipFile('resources/archive.zip', 'r') as archive:
        with archive.open('testfile1.pdf') as pdffile:
            content_pdf = pdffile.read()
            pdf_stream = io.BytesIO(content_pdf)
            reader = PdfReader(pdf_stream)
            content = ""
            for page in reader.pages:
                content += page.extract_text()
                expected_value = 'Lorem ipsum'
                assert expected_value in content
                print(f"Содержимое файла PDF проверено успешно. Слова: '{expected_value}' найдены в файле")
        with archive.open('testfile2.xlsx') as xlsxfile:
            content_xlsx = xlsxfile.read()
            xlsx_stream = io.BytesIO(content_xlsx)
            workbook = load_workbook(xlsx_stream)
            sheet = workbook.active
            cell_value = sheet.cell(row=1, column=1).value
            expected_value = 'Postcode'
            assert cell_value == expected_value
            print(f"Содержимое файла XLSX проверено успешно. Значение в ячейке: '{cell_value}'")
        with archive.open('testfile3.csv') as csvfile:
            content_csv = csvfile.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content_csv.splitlines()))
            second_row = csvreader[1]
            assert second_row[0] == 'BDCQ.SEA1AA'
            assert second_row[1] == '2011.06'
            print(f"Содержимое файла CSV проверено успешно. Значение в ячейке 1: '{second_row[0]}', значение в ячейке 2:{second_row[1]}'")
